import calendar as pycal
from datetime import datetime
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

# =========================================================
# CONFIG
# =========================================================
APP_TITLE = "dpkFXdashboard"
st.set_page_config(page_title=APP_TITLE, page_icon="📈", layout="wide")

MONTH_MAP = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
             'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}

EXPECTED_COLS = ['Blank', '#', 'Mth', 'P', 'Date', 'WK', 'W/L', 'Pair', 'Action',
                  'Pip', 'Session', 'Lot', 'Swap', 'Charges', 'P&L', 'Day Total',
                  'Actl Pft %', 'TP', 'SL', 'RR', 'R Multiple', 'Risk %', 'Exit',
                  'Actual Hit', 'Why', 'Setup', 'Model', 'EN/T', 'Avd']

# =========================================================
# THEME — light cards on light-gray bg, purple accent,
# mint green / coral red for P&L
# =========================================================
THEMES = {
    "Light": dict(
        bg="#f4f5fa", card_bg="#ffffff", border="#eceef5",
        text_primary="#161722", text_secondary="#8a8fa3", ticker_bg="#ffffff",
        green="#1fce8f", green_soft="rgba(31,206,143,0.18)",
        red="#f2637d", red_soft="rgba(242,99,125,0.18)",
        purple="#7b61ff", purple_soft="rgba(123,97,255,0.12)",
        plotly_template="plotly_white", plot_bg="#ffffff", shadow="0 1px 3px rgba(20,20,40,0.06)",
    ),
    "Dark": dict(
        bg="#0d0e1a", card_bg="#16172a", border="#242540",
        text_primary="#f1f2f6", text_secondary="#8b8fa3", ticker_bg="#16172a",
        green="#1fce8f", green_soft="rgba(31,206,143,0.18)",
        red="#f2637d", red_soft="rgba(242,99,125,0.18)",
        purple="#8b5cf6", purple_soft="rgba(139,92,246,0.18)",
        plotly_template="plotly_dark", plot_bg="#16172a", shadow="0 1px 3px rgba(0,0,0,0.4)",
    ),
}
if "theme" not in st.session_state:
    st.session_state.theme = "Light"
T = THEMES[st.session_state.theme]

st.markdown(f"""
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stHeader"],
    [data-testid="stToolbar"], .main, .block-container {{
        background-color: {T['bg']} !important; color: {T['text_primary']} !important;
        font-family: 'Inter', sans-serif !important;
    }}
    h1, h2, h3, h4, p, span, label, div {{ color: {T['text_primary']}; font-family: 'Inter', sans-serif; }}
    h1 {{ font-weight: 800; letter-spacing: -0.5px; }}
    h3, h4 {{ font-weight: 700; }}
    .card {{ background: {T['card_bg']}; border: 1px solid {T['border']}; border-radius: 14px;
             padding: 16px 18px; box-shadow: {T['shadow']}; margin-bottom: 12px; }}
    .kpi-label {{ color: {T['text_secondary']}; font-size: 12px; font-weight: 600; margin-bottom: 6px; }}
    .kpi-value {{ font-size: 24px; font-weight: 800; }}
    .kpi-sub {{ font-size: 11.5px; color: {T['text_secondary']}; margin-top: 2px; }}
    .green {{ color: {T['green']}; }} .red {{ color: {T['red']}; }} .purple {{ color: {T['purple']}; }}
    .neutral {{ color: {T['text_primary']}; }}
    .ticker-wrap {{ width: 100%; overflow: hidden; background: {T['ticker_bg']}; border: 1px solid {T['border']};
                     border-radius: 12px; padding: 10px 0; margin-bottom: 18px; box-shadow: {T['shadow']}; }}
    .ticker-move {{ display: inline-block; white-space: nowrap; animation: ticker 140s linear infinite; font-size: 13px; font-weight: 500; }}
    .ticker-wrap:hover .ticker-move {{ animation-play-state: paused; }}
    @keyframes ticker {{ 0% {{ transform: translateX(0); }} 100% {{ transform: translateX(-50%); }} }}
    .ticker-item {{ display: inline-block; padding: 0 24px; }}
    div[data-testid="stDataFrame"] {{ border: 1px solid {T['border']}; border-radius: 12px; }}
    div[data-testid="stExpander"] {{ background: {T['card_bg']}; border: 1px solid {T['border']}; border-radius: 14px; }}
    .stButton>button {{ background: {T['purple']} !important; color: white !important; border: none !important;
                         border-radius: 8px !important; font-weight: 600 !important; }}
    .cal-card {{ background: {T['card_bg']}; border: 1px solid {T['border']}; border-radius: 14px; padding: 14px; box-shadow: {T['shadow']}; }}
    .cal-grid {{ display: grid; grid-template-columns: repeat(7, 1fr); gap: 6px; }}
    .cal-head {{ color: {T['text_secondary']}; font-size: 11px; font-weight: 700; text-transform: uppercase; text-align: center; padding: 4px 0; }}
    .cal-cell {{ border-radius: 8px; padding: 6px 7px; min-height: 74px; font-size: 11px; }}
    .cal-cell .d {{ font-size: 11px; opacity: 0.55; }}
    .cal-cell .pnl {{ font-size: 14px; font-weight: 800; margin-top: 5px; }}
    .cal-cell .cnt {{ font-size: 10px; opacity: 0.8; margin-top: 3px; }}
    .cal-cell .r {{ font-size: 10px; opacity: 0.8; }}
    .cal-empty {{ background: transparent; }}
    .week-card {{ background: {T['card_bg']}; border: 1px solid {T['border']}; border-radius: 10px; padding: 10px 12px; margin-bottom: 8px; box-shadow: {T['shadow']}; }}
    .week-label {{ font-size: 11px; color: {T['text_secondary']}; font-weight: 600; }}
    .week-val {{ font-size: 15px; font-weight: 800; margin-top: 2px; }}
    .week-days {{ font-size: 10.5px; color: {T['text_secondary']}; margin-top: 2px; }}
    .tag-row {{ display: flex; justify-content: space-between; padding: 8px 4px; border-bottom: 1px solid {T['border']}; font-size: 13px; }}
</style>
""", unsafe_allow_html=True)


def themed(fig, height=None, margin=None):
    fig.update_layout(template=T['plotly_template'], paper_bgcolor=T['card_bg'],
                       plot_bgcolor=T['card_bg'], font=dict(color=T['text_primary'], family='Inter'))
    if height:
        fig.update_layout(height=height)
    if margin:
        fig.update_layout(margin=margin)
    return fig


# =========================================================
# DATA LOADING & PARSING (Excel upload)
# =========================================================
def clean_numeric(series):
    return pd.to_numeric(
        series.astype(str).str.replace('%', '', regex=False).str.replace(',', '', regex=False)
        .str.replace('₹', '', regex=False).str.strip()
        .replace({'-': None, '': None, 'nan': None, 'None': None, '#DIV/0!': None, '#N/A': None, '#REF!': None}),
        errors='coerce'
    )


def parse_date(val, mth_hint):
    """Handles both text 'dd/mm/yyyy' cells and native Excel date cells.
    Native date cells can have day/month swapped when Excel auto-parsed an
    ambiguous dd/mm entry under a different locale — corrected using the
    sheet's own 'Mth' column as a cross-check."""
    if pd.isna(val) or val == '':
        return pd.NaT
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), '%d/%m/%Y')
        except Exception:
            return pd.NaT
    if isinstance(val, datetime):
        hint = MONTH_MAP.get(str(mth_hint).strip().lower()[:3]) if pd.notna(mth_hint) else None
        swapped = None
        if val.day <= 12:
            try:
                swapped = datetime(val.year, val.day, val.month)
            except ValueError:
                swapped = None
        if hint is not None:
            if swapped is not None and swapped.month == hint:
                return swapped
            if val.month == hint:
                return val
        return swapped if swapped is not None else val
    return pd.NaT


@st.cache_data(show_spinner=False)
def load_excel(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, min(ws.max_row, 100) + 1):
        v1 = ws.cell(row=r, column=2).value
        v4 = ws.cell(row=r, column=5).value
        if str(v1).strip() == '#' and str(v4).strip() == 'Date':
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not locate the trade-journal header row in this sheet.")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 30)]
        rows.append(row)

    df = pd.DataFrame(rows, columns=EXPECTED_COLS).drop(columns=['Blank'])

    df['Date'] = [parse_date(v, m) for v, m in zip(df['Date'], df['Mth'])]
    df = df[df['Date'].notna()]
    df = df[df['W/L'].isin(['Won', 'Loss'])]

    for col in ['Pip', 'Lot', 'Swap', 'Charges', 'P&L', 'Day Total', 'Actl Pft %',
                'TP', 'SL', 'RR', 'R Multiple']:
        df[col] = clean_numeric(df[col])

    df['Win'] = (df['W/L'] == 'Won').astype(int)
    df['Setup'] = df['Setup'].fillna('Unknown').replace('', 'Unknown').str.strip()
    df['Model'] = df['Model'].fillna('N/A').replace('', 'N/A').str.strip()
    df['Model'] = df['Model'].replace({'Volume/E...': 'Volume / Engulf', 'Volume/E': 'Volume / Engulf'})
    df['Session'] = df['Session'].fillna('Unknown').replace('', 'Unknown').str.strip()
    df['Pair'] = df['Pair'].fillna('Unknown').replace('', 'Unknown').str.strip()
    df['Action'] = df['Action'].fillna('Unknown').replace('', 'Unknown').str.strip()
    df['WK'] = df['WK'].fillna('N/A').replace('', 'N/A')
    df['Why'] = df['Why'].fillna('').astype(str)
    df['Month'] = df['Date'].dt.to_period('M').dt.to_timestamp()

    df = df.sort_values('Date').reset_index(drop=True)
    return df


# =========================================================
# HEADER + FILE UPLOAD
# =========================================================
hour = datetime.now().hour
greeting = "Good morning" if hour < 12 else ("Good afternoon" if hour < 18 else "Good evening")

h1, h2 = st.columns([5, 1])
with h1:
    st.markdown(f"# 📈 {APP_TITLE}")
    st.caption(f"{greeting}! Here's how your trading is going.")
with h2:
    choice = st.radio("Theme", ["Light", "Dark"], horizontal=True,
                       index=0 if st.session_state.theme == "Light" else 1, label_visibility="collapsed")
    if choice != st.session_state.theme:
        st.session_state.theme = choice
        st.rerun()

uploaded = st.file_uploader("Upload your trading journal (.xlsx)", type=["xlsx"])
if uploaded is None:
    st.info("Upload your Trading_Journal .xlsx file above to load the dashboard.")
    st.stop()

file_bytes = uploaded.getvalue()
wb_peek = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
sheet_names = wb_peek.sheetnames
default_sheet = 'CRT26' if 'CRT26' in sheet_names else sheet_names[0]
sel_sheet = st.selectbox("Sheet", sheet_names, index=sheet_names.index(default_sheet))

try:
    df_raw = load_excel(file_bytes, sel_sheet)
except Exception as e:
    st.error(f"Could not parse this sheet: {e}")
    st.stop()

if df_raw.empty:
    st.warning("No trade rows were found in this sheet.")
    st.stop()

# =========================================================
# FILTERS
# =========================================================
with st.expander("⚙️ Filters", expanded=False):
    min_date, max_date = df_raw['Date'].min().date(), df_raw['Date'].max().date()
    date_range = st.date_input("Date range", value=(min_date, max_date), min_value=min_date, max_value=max_date)
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
    else:
        start_date, end_date = min_date, max_date

    fr1, fr2, fr3, fr4 = st.columns(4)
    with fr1:
        pairs = sorted(df_raw['Pair'].unique())
        sel_pairs = st.multiselect("Currency Pair", pairs, default=pairs)
        models = sorted(df_raw['Model'].unique())
        sel_models = st.multiselect("Setup / Model", models, default=models)
    with fr2:
        setups = sorted(df_raw['Setup'].unique())
        sel_setups = st.multiselect("Trade Type", setups, default=setups)
        sessions = sorted(df_raw['Session'].unique())
        sel_sessions = st.multiselect("Session", sessions, default=sessions)
    with fr3:
        actions = sorted(df_raw['Action'].unique())
        sel_actions = st.multiselect("Action", actions, default=actions)
        wl_options = ['Won', 'Loss']
        sel_wl = st.multiselect("Result", wl_options, default=wl_options)
    with fr4:
        st.caption(f"Latest data point: **{max_date}**")
        st.caption(f"Total rows loaded: **{len(df_raw)}**")

mask = (
    (df_raw['Date'].dt.date >= start_date) & (df_raw['Date'].dt.date <= end_date) &
    (df_raw['Pair'].isin(sel_pairs)) & (df_raw['Model'].isin(sel_models)) &
    (df_raw['Setup'].isin(sel_setups)) & (df_raw['Session'].isin(sel_sessions)) &
    (df_raw['Action'].isin(sel_actions)) & (df_raw['W/L'].isin(sel_wl))
)
df = df_raw[mask].copy()
if df.empty:
    st.warning("No trades match the current filters.")
    st.stop()

df = df.sort_values('Date').reset_index(drop=True)
df['Cum P&L'] = df['P&L'].fillna(0).cumsum()
df['Running Max'] = df['Cum P&L'].cummax()
df['Drawdown'] = df['Cum P&L'] - df['Running Max']
df['Cum R'] = df['R Multiple'].fillna(0).cumsum()
df['Running Max R'] = df['Cum R'].cummax()

# =========================================================
# TICKER
# =========================================================
ticker_items = ""
for _, r in df.sort_values('Date', ascending=False).head(30).iterrows():
    color = T['green'] if r['W/L'] == 'Won' else T['red']
    sign = "+" if (r['P&L'] or 0) >= 0 else ""
    pnl_txt = f"{sign}{r['P&L']:.2f}" if pd.notna(r['P&L']) else "—"
    ticker_items += (f"<span class='ticker-item'>{r['Date'].strftime('%d %b')} · <b>{r['Pair']}</b> {r['Action']} "
                      f"<span style='color:{color}'>{pnl_txt}</span></span>")
st.markdown(f"<div class='ticker-wrap'><div class='ticker-move'>{ticker_items}{ticker_items}</div></div>",
            unsafe_allow_html=True)

# =========================================================
# CORE METRICS
# =========================================================
total_trades = len(df)
wins = int(df['Win'].sum())
losses = total_trades - wins
win_rate = (wins / total_trades * 100) if total_trades else 0
total_pnl = df['P&L'].sum()
avg_rr = df['R Multiple'].mean()
gross_profit = df.loc[df['P&L'] > 0, 'P&L'].sum()
gross_loss = abs(df.loc[df['P&L'] < 0, 'P&L'].sum())
profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else np.nan
avg_win = df.loc[df['Win'] == 1, 'P&L'].mean() if wins else 0
avg_loss = df.loc[df['Win'] == 0, 'P&L'].mean() if losses else 0
expectancy = total_pnl / total_trades if total_trades else 0
max_dd = df['Drawdown'].min()


def streak_len(series, from_end=True):
    seq = series.iloc[::-1] if from_end else series
    s, s_type = 0, None
    for w in seq:
        cur = 'W' if w == 1 else 'L'
        if s_type is None:
            s_type, s = cur, 1
        elif cur == s_type:
            s += 1
        else:
            break
    return s, s_type


def max_streak(series, target):
    m = c = 0
    for x in series:
        if x == target:
            c += 1
            m = max(m, c)
        else:
            c = 0
    return m


streak, streak_type = streak_len(df['Win'])
longest_win_streak = max_streak(df['Win'], 1)
longest_loss_streak = max_streak(df['Win'], 0)

# =========================================================
# KPI ROW
# =========================================================
k1, k2, k3, k4, k5, k6 = st.columns(6)
with k1:
    cls = "green" if total_pnl >= 0 else "red"
    sign = "+" if total_pnl >= 0 else ""
    st.markdown(f"""<div class="card"><div class="kpi-label">NET P&L</div>
        <div class="kpi-value {cls}">{sign}{total_pnl:,.2f}</div>
        <div class="kpi-sub">{total_trades} trades</div></div>""", unsafe_allow_html=True)

with k2:
    st.markdown('<div class="card"><div class="kpi-label">PROFIT FACTOR</div>', unsafe_allow_html=True)
    fig_pf = go.Figure(go.Indicator(
        mode="gauge+number", value=0 if pd.isna(profit_factor) else profit_factor,
        number={'font': {'size': 20, 'color': T['text_primary']}, 'valueformat': '.2f'},
        gauge={'shape': "angular", 'axis': {'range': [0, 3], 'visible': False},
               'bar': {'color': T['purple'], 'thickness': 0.3}, 'bgcolor': T['card_bg'], 'borderwidth': 0,
               'steps': [{'range': [0, 1], 'color': T['red_soft']}, {'range': [1, 3], 'color': T['green_soft']}]}))
    themed(fig_pf, height=110, margin=dict(l=10, r=10, t=0, b=0))
    st.plotly_chart(fig_pf, use_container_width=True, config={'displayModeBar': False})
    st.markdown('</div>', unsafe_allow_html=True)

with k3:
    st.markdown('<div class="card"><div class="kpi-label">TRADE WIN %</div>', unsafe_allow_html=True)
    fig_wr = go.Figure(go.Pie(values=[wins, losses], hole=0.72, marker_colors=[T['green'], T['red']],
                               textinfo='none', sort=False))
    fig_wr.add_annotation(text=f"<b>{win_rate:.1f}%</b>", showarrow=False, font=dict(size=17, color=T['text_primary']))
    themed(fig_wr, height=110, margin=dict(l=10, r=10, t=0, b=0))
    fig_wr.update_layout(showlegend=False)
    st.plotly_chart(fig_wr, use_container_width=True, config={'displayModeBar': False})
    st.markdown(f"<div class='kpi-sub'><span class='green'>{wins}W</span> / <span class='red'>{losses}L</span></div></div>",
                unsafe_allow_html=True)

with k4:
    st.markdown('<div class="card"><div class="kpi-label">AVG WIN / LOSS</div>', unsafe_allow_html=True)
    fig_avgwl = go.Figure()
    fig_avgwl.add_trace(go.Bar(x=[avg_win], y=[''], orientation='h', marker_color=T['green'], showlegend=False))
    fig_avgwl.add_trace(go.Bar(x=[avg_loss], y=[''], orientation='h', marker_color=T['red'], showlegend=False))
    themed(fig_avgwl, height=70, margin=dict(l=10, r=10, t=10, b=0))
    fig_avgwl.update_layout(barmode='overlay', xaxis=dict(visible=False), yaxis=dict(visible=False))
    st.plotly_chart(fig_avgwl, use_container_width=True, config={'displayModeBar': False})
    st.markdown(f"<div class='kpi-sub'><span class='green'>{avg_win:,.0f}</span> / <span class='red'>{avg_loss:,.0f}</span></div></div>",
                unsafe_allow_html=True)

with k5:
    cls = "green" if expectancy >= 0 else "red"
    st.markdown(f"""<div class="card"><div class="kpi-label">TRADE EXPECTANCY</div>
        <div class="kpi-value {cls}">{expectancy:,.2f}</div><div class="kpi-sub">avg P&L per trade</div></div>""",
        unsafe_allow_html=True)

with k6:
    cls = "green" if streak_type == 'W' else "red"
    st.markdown(f"""<div class="card"><div class="kpi-label">CURRENT STREAK</div>
        <div class="kpi-value {cls}">{streak}{streak_type or ''}</div>
        <div class="kpi-sub">longest: {longest_win_streak}W / {longest_loss_streak}L</div></div>""",
        unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# =========================================================
# ROW 2 — Performance Score | Daily Net Cumulative P&L | Net Daily P&L
# =========================================================
r1, r2, r3 = st.columns([1, 1.4, 1.4])
with r1:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("**Performance Score**")
    win_score = win_rate
    ratio = (avg_win / abs(avg_loss)) if avg_loss else 0
    avgwl_score = min(ratio / 3, 1) * 100 if ratio > 0 else 0
    pf_score = min(profit_factor / 3, 1) * 100 if pd.notna(profit_factor) else 0
    overall_score = np.mean([win_score, avgwl_score, pf_score])
    cats = ['Win %', 'Avg Win/Loss', 'Profit Factor']
    vals = [win_score, avgwl_score, pf_score]
    fig_radar = go.Figure(go.Scatterpolar(r=vals + [vals[0]], theta=cats + [cats[0]], fill='toself',
                                           line_color=T['purple'], fillcolor=T['purple_soft']))
    themed(fig_radar, height=230, margin=dict(l=30, r=30, t=10, b=10))
    fig_radar.update_layout(polar=dict(bgcolor=T['card_bg'],
                             radialaxis=dict(visible=True, range=[0, 100], showticklabels=False, gridcolor=T['border']),
                             angularaxis=dict(gridcolor=T['border'])), showlegend=False)
    st.plotly_chart(fig_radar, use_container_width=True, config={'displayModeBar': False})
    st.markdown(f"<div style='text-align:center'>Your Score: <span class='purple' style='font-weight:800;font-size:18px'>{overall_score:.1f}</span></div>",
                unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

with r2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("**Daily Net Cumulative P&L**")
    is_high = df['Cum P&L'] >= df['Running Max']
    green_y = df['Cum P&L'].where(is_high)
    red_y = df['Cum P&L'].where(~is_high)
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(x=df['Date'], y=green_y, mode='lines', line=dict(color=T['green'], width=2),
                                  fill='tozeroy', fillcolor=T['green_soft'], connectgaps=False))
    fig_cum.add_trace(go.Scatter(x=df['Date'], y=red_y, mode='lines', line=dict(color=T['red'], width=2),
                                  fill='tozeroy', fillcolor=T['red_soft'], connectgaps=False))
    themed(fig_cum, height=270, margin=dict(l=10, r=10, t=10, b=10))
    fig_cum.update_layout(showlegend=False)
    st.plotly_chart(fig_cum, use_container_width=True, config={'displayModeBar': False})
    st.markdown('</div>', unsafe_allow_html=True)

with r3:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("**Net Daily P&L**")
    daily_pnl_series = df.groupby(df['Date'].dt.date)['P&L'].sum()
    fig_daily = go.Figure(go.Bar(x=[d.strftime('%d %b') for d in daily_pnl_series.index], y=daily_pnl_series.values,
                                  marker_color=[T['green'] if v >= 0 else T['red'] for v in daily_pnl_series.values]))
    themed(fig_daily, height=270, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig_daily, use_container_width=True, config={'displayModeBar': False})
    st.markdown('</div>', unsafe_allow_html=True)

# =========================================================
# R-MULTIPLE EQUITY CURVE — sizing-independent performance
# =========================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown("**R-Multiple Equity Curve** — cumulative R, independent of lot size")
is_high_r = df['Cum R'] >= df['Running Max R']
green_r = df['Cum R'].where(is_high_r)
red_r = df['Cum R'].where(~is_high_r)
fig_r_eq = go.Figure()
fig_r_eq.add_trace(go.Scatter(x=df['Date'], y=green_r, mode='lines', line=dict(color=T['purple'], width=2),
                               fill='tozeroy', fillcolor=T['purple_soft'], connectgaps=False))
fig_r_eq.add_trace(go.Scatter(x=df['Date'], y=red_r, mode='lines', line=dict(color=T['red'], width=2),
                               fill='tozeroy', fillcolor=T['red_soft'], connectgaps=False))
themed(fig_r_eq, height=260, margin=dict(l=10, r=10, t=10, b=10))
fig_r_eq.update_layout(showlegend=False, yaxis_title="Cumulative R")
st.plotly_chart(fig_r_eq, use_container_width=True, config={'displayModeBar': False})
st.markdown('</div>', unsafe_allow_html=True)

# =========================================================
# CALENDAR
# =========================================================
st.markdown("### Calendar")
daily = df.groupby(df['Date'].dt.normalize()).agg(
    Trades=('P&L', 'count'), PnL=('P&L', 'sum'), Wins=('Win', 'sum'), R=('R Multiple', 'sum')
).reset_index().rename(columns={'Date': 'Day'})
daily_map = {row['Day'].date(): row for _, row in daily.iterrows()}

available_months = sorted(df['Date'].dt.to_period('M').unique())
month_labels = [m.strftime('%B %Y') for m in available_months]
if "cal_month_idx" not in st.session_state or st.session_state.cal_month_idx >= len(available_months):
    st.session_state.cal_month_idx = len(available_months) - 1

nav1, nav2, nav3 = st.columns([1, 3, 1])
with nav1:
    if st.button("◀ Prev") and st.session_state.cal_month_idx > 0:
        st.session_state.cal_month_idx -= 1
        st.rerun()
with nav3:
    if st.button("Next ▶") and st.session_state.cal_month_idx < len(available_months) - 1:
        st.session_state.cal_month_idx += 1
        st.rerun()
with nav2:
    st.markdown(f"<h4 style='text-align:center'>{month_labels[st.session_state.cal_month_idx]}</h4>", unsafe_allow_html=True)

sel_period = available_months[st.session_state.cal_month_idx]
year, month = sel_period.year, sel_period.month
cal_obj = pycal.Calendar(firstweekday=6)
month_weeks = cal_obj.monthdatescalendar(year, month)
max_abs_day = max([abs(v['PnL']) for v in daily_map.values()] + [1])


def cell_style(pnl):
    if pnl is None:
        return "background: transparent;"
    intensity = min(abs(pnl) / max_abs_day, 1) * 0.35 + 0.10
    color = T['green'] if pnl >= 0 else T['red']
    rgb = "31,206,143" if pnl >= 0 else "242,99,125"
    return f"background: rgba({rgb},{intensity}); border: 1px solid {color};"


cal_col, week_col = st.columns([3, 1])
with cal_col:
    html = "<div class='cal-card'><div class='cal-grid'>"
    for wd in ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']:
        html += f"<div class='cal-head'>{wd}</div>"
    for week in month_weeks:
        for d in week:
            if d.month != month:
                html += "<div class='cal-cell cal-empty'></div>"
                continue
            row = daily_map.get(d)
            if row is not None and row['Trades'] > 0:
                pnl, trades, wins_d, r_d = row['PnL'], int(row['Trades']), int(row['Wins']), row['R']
                day_wr = (wins_d / trades * 100) if trades else 0
                color_cls = 'green' if pnl >= 0 else 'red'
                r_txt = f"{r_d:.2f}R, " if pd.notna(r_d) else ""
                html += (f"<div class='cal-cell' style='{cell_style(pnl)}'><div class='d'>{d.day}</div>"
                          f"<div class='pnl {color_cls}'>{pnl:+,.0f}</div>"
                          f"<div class='cnt'>{trades} trade{'s' if trades != 1 else ''}</div>"
                          f"<div class='r'>{r_txt}{day_wr:.0f}%</div></div>")
            else:
                html += f"<div class='cal-cell' style='{cell_style(None)}'><div class='d'>{d.day}</div></div>"
    html += "</div></div>"
    st.markdown(html, unsafe_allow_html=True)

with week_col:
    for i, week in enumerate(month_weeks, start=1):
        week_pnl, week_days = 0.0, 0
        for d in week:
            if d.month != month:
                continue
            row = daily_map.get(d)
            if row is not None and row['Trades'] > 0:
                week_pnl += row['PnL']
                week_days += 1
        if all(d.month != month for d in week):
            continue
        cls = 'green' if week_pnl >= 0 else 'red'
        st.markdown(f"""<div class="week-card"><div class="week-label">Week {i}</div>
            <div class="week-val {cls}">{week_pnl:+,.2f}</div>
            <div class="week-days">{week_days} day{'s' if week_days != 1 else ''}</div></div>""",
            unsafe_allow_html=True)

# =========================================================
# WEEK-WISE PERFORMANCE
# =========================================================
st.markdown("### Week-wise Performance")
week_order = df.groupby('WK')['Date'].min().sort_values().index.tolist()
weekly = df.groupby('WK').agg(Trades=('P&L', 'count'), Wins=('Win', 'sum'), PnL=('P&L', 'sum'),
                               WeekStart=('Date', 'min'), WeekEnd=('Date', 'max')).reindex(week_order).reset_index()
weekly['Win Rate %'] = (weekly['Wins'] / weekly['Trades'] * 100).round(1)

fig_week = go.Figure()
fig_week.add_trace(go.Bar(x=weekly['WK'], y=weekly['PnL'], name='P&L',
                           marker_color=[T['green'] if v >= 0 else T['red'] for v in weekly['PnL']], yaxis='y1'))
fig_week.add_trace(go.Scatter(x=weekly['WK'], y=weekly['Trades'], name='Trades', mode='lines+markers',
                               line=dict(color=T['purple'], width=2), yaxis='y2'))
themed(fig_week, height=340, margin=dict(l=10, r=10, t=30, b=10))
fig_week.update_layout(yaxis=dict(title='P&L'), yaxis2=dict(title='Trades', overlaying='y', side='right', showgrid=False),
                        legend=dict(orientation='h', y=1.1))
st.markdown('<div class="card">', unsafe_allow_html=True)
st.plotly_chart(fig_week, use_container_width=True, config={'displayModeBar': False})
st.markdown('</div>', unsafe_allow_html=True)

with st.expander("Week-wise data table"):
    wk_table = weekly.copy()
    wk_table['WeekStart'] = wk_table['WeekStart'].dt.strftime('%d %b %Y')
    wk_table['WeekEnd'] = wk_table['WeekEnd'].dt.strftime('%d %b %Y')
    st.dataframe(wk_table[['WK', 'WeekStart', 'WeekEnd', 'Trades', 'Wins', 'Win Rate %', 'PnL']]
                 .rename(columns={'PnL': 'P&L'}).style.format({'P&L': '{:,.2f}', 'Win Rate %': '{:.1f}'}),
                 use_container_width=True, hide_index=True)

# =========================================================
# BEHAVIOR / MISTAKE TAG ANALYSIS (from the "Why?" column)
# =========================================================
st.markdown("### Behavior & Mistake Analysis")
st.caption("Parsed from your 'Why?' notes — which habits are actually costing or making you money.")

exploded = df[['Why', 'P&L', 'Win']].copy()
exploded = exploded[exploded['Why'].str.strip() != '']
exploded['Why'] = exploded['Why'].str.split(',')
exploded = exploded.explode('Why')
exploded['Why'] = exploded['Why'].str.strip().str.title()
exploded = exploded[exploded['Why'] != '']

if not exploded.empty:
    tag_stats = exploded.groupby('Why').agg(Count=('P&L', 'count'), TotalPnL=('P&L', 'sum'),
                                              WinRate=('Win', 'mean')).sort_values('Count', ascending=False).head(15)
    tag_stats['WinRate'] = (tag_stats['WinRate'] * 100).round(1)

    bc1, bc2 = st.columns([1.3, 1])
    with bc1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**Most frequent tags — $ impact**")
        tag_sorted = tag_stats.sort_values('TotalPnL')
        fig_tags = go.Figure(go.Bar(
            x=tag_sorted['TotalPnL'], y=tag_sorted.index, orientation='h',
            marker_color=[T['green'] if v >= 0 else T['red'] for v in tag_sorted['TotalPnL']],
            text=tag_sorted['Count'].astype(str) + 'x', textposition='outside',
        ))
        themed(fig_tags, height=420, margin=dict(l=10, r=10, t=10, b=10))
        fig_tags.update_layout(xaxis_title="Total P&L impact")
        st.plotly_chart(fig_tags, use_container_width=True, config={'displayModeBar': False})
        st.markdown('</div>', unsafe_allow_html=True)
    with bc2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**Tag breakdown**")
        st.dataframe(
            tag_stats.rename(columns={'TotalPnL': 'Total P&L', 'WinRate': 'Win Rate %'})
            .style.format({'Total P&L': '{:,.2f}', 'Win Rate %': '{:.1f}'}),
            use_container_width=True, height=420,
        )
        st.markdown('</div>', unsafe_allow_html=True)
else:
    st.info("No 'Why?' notes found in the filtered trades.")

# =========================================================
# PERFORMANCE BREAKDOWNS — Session | Pair | Action | My vs Not My Setup
# =========================================================
st.markdown("### Performance Breakdowns")
bd1, bd2, bd3, bd4 = st.columns(4)


def breakdown_chart(col, group_col, title):
    stats = df.groupby(group_col).agg(Trades=('P&L', 'count'), PnL=('P&L', 'sum'), Wins=('Win', 'sum')).reset_index()
    stats['WR'] = (stats['Wins'] / stats['Trades'] * 100).round(0)
    stats = stats.sort_values('PnL', ascending=True)
    fig = go.Figure(go.Bar(x=stats['PnL'], y=stats[group_col], orientation='h',
                            marker_color=[T['green'] if v >= 0 else T['red'] for v in stats['PnL']],
                            text=stats['WR'].astype(int).astype(str) + '%', textposition='outside'))
    themed(fig, height=280, margin=dict(l=10, r=10, t=25, b=10))
    with col:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f"**{title}**")
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        st.markdown('</div>', unsafe_allow_html=True)


breakdown_chart(bd1, 'Session', 'By Session')
breakdown_chart(bd2, 'Pair', 'By Pair')
breakdown_chart(bd3, 'Action', 'Buy vs Sell')
breakdown_chart(bd4, 'Setup', 'Planned vs Impulsive')

# =========================================================
# SETUP / MODEL PERFORMANCE + R-MULTIPLE DISTRIBUTION
# =========================================================
col_left2, col_right2 = st.columns(2)
with col_left2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("**Setup Performance (by Model)**")
    setup_perf = df.groupby('Model').agg(Trades=('P&L', 'count'), Wins=('Win', 'sum'), Total_PnL=('P&L', 'sum'),
                                          Avg_RR=('R Multiple', 'mean')).reset_index()
    setup_perf['Win Rate %'] = (setup_perf['Wins'] / setup_perf['Trades'] * 100).round(1)
    setup_perf = setup_perf.sort_values('Total_PnL', ascending=False)
    fig_setup = go.Figure(go.Bar(x=setup_perf['Model'], y=setup_perf['Total_PnL'],
                                  marker_color=[T['green'] if v >= 0 else T['red'] for v in setup_perf['Total_PnL']],
                                  text=setup_perf['Win Rate %'].astype(str) + '% WR', textposition='outside'))
    themed(fig_setup, height=300, margin=dict(l=10, r=10, t=20, b=10))
    st.plotly_chart(fig_setup, use_container_width=True, config={'displayModeBar': False})
    st.markdown('</div>', unsafe_allow_html=True)

with col_right2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("**R-Multiple Distribution**")
    r_data = df['R Multiple'].dropna()
    if len(r_data):
        fig_r = go.Figure(go.Histogram(x=r_data, nbinsx=30, marker_color=T['purple']))
        themed(fig_r, height=300, margin=dict(l=10, r=10, t=10, b=10))
        fig_r.add_vline(x=0, line_dash="dash", line_color=T['text_secondary'])
        st.plotly_chart(fig_r, use_container_width=True, config={'displayModeBar': False})
    else:
        st.info("No R-multiple data available.")
    st.markdown('</div>', unsafe_allow_html=True)

# =========================================================
# BEST / WORST TRADES
# =========================================================
st.markdown("### Best & Worst Trades")
bw1, bw2 = st.columns(2)
best = df.loc[df['P&L'].idxmax()]
worst = df.loc[df['P&L'].idxmin()]
with bw1:
    st.markdown(f"""<div class="card">
        <div class="kpi-label">BEST TRADE</div>
        <div class="kpi-value green">+{best['P&L']:,.2f}</div>
        <div class="kpi-sub">{best['Date'].strftime('%d %b %Y')} · {best['Pair']} {best['Action']} · {best['Model']}</div>
    </div>""", unsafe_allow_html=True)
with bw2:
    st.markdown(f"""<div class="card">
        <div class="kpi-label">WORST TRADE</div>
        <div class="kpi-value red">{worst['P&L']:,.2f}</div>
        <div class="kpi-sub">{worst['Date'].strftime('%d %b %Y')} · {worst['Pair']} {worst['Action']} · {worst['Model']}</div>
    </div>""", unsafe_allow_html=True)

# =========================================================
# RECENT TRADES TABLE
# =========================================================
st.markdown("### Recent Trades")
show_cols = ['Date', 'Pair', 'Action', 'Session', 'W/L', 'Lot', 'Pip', 'P&L',
             'R Multiple', 'RR', 'Setup', 'Model', 'Why']
display_df = df.sort_values('Date', ascending=False)[show_cols].copy()
display_df['Date'] = display_df['Date'].dt.strftime('%d %b %Y')
st.dataframe(
    display_df.style.format({'P&L': '{:,.2f}', 'R Multiple': '{:.2f}', 'RR': '{:.2f}', 'Pip': '{:.1f}'})
    .map(lambda v: f"color: {T['green']}" if v == 'Won' else (f"color: {T['red']}" if v == 'Loss' else ''), subset=['W/L']),
    use_container_width=True, hide_index=True, height=420,
)

st.caption(f"Showing {total_trades} trades matching current filters.")
